# -*- coding: utf-8 -*-
#!/usr/bin/env python3
import argparse
import os
import re
import requests
import pandas as pd
import colorsys
from openpyxl import load_workbook
from urllib.parse import urlparse, unquote
import ast
import sys
import socket
import getpass
from pathlib import Path

# =========================
# Color / ordinal helpers
# =========================
from openpyxl import load_workbook
import pandas as pd, re, colorsys
from typing import Dict, Tuple

def rgb_to_signed_score(rgb_hex: str | None) -> float:
    if not rgb_hex:
        return 0.0
    if len(rgb_hex) == 8:  # 'FFRRGGBB' -> 'RRGGBB'
        rgb_hex = rgb_hex[2:]
    r = int(rgb_hex[0:2], 16)/255
    g = int(rgb_hex[2:4], 16)/255
    b = int(rgb_hex[4:6], 16)/255
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    hue = h*360
    inten = s*v
    if 90 <= hue <= 150:       # green band
        return +float(inten)
    if hue <= 15 or hue >= 345:  # red band
        return -float(inten)
    return 0.0  # neutral/other hues

def signed_score_to_likert(score: float, n_levels: int = 5) -> int:
    if score == 0:
        return 0
    a = min(abs(score), 1.0)
    lvl = max(1, round(a * n_levels))
    return lvl if score > 0 else -lvl

def theme_to_rgb(theme_index: int | None, theme_map: Dict[int, str] | None = None) -> str | None:
    base = {9: 'FF00FF00', 10: 'FFFF0000'}  # extend if needed
    if theme_map:
        base.update(theme_map)
    return base.get(theme_index, None)

def extract_third_sheet_multiheader_and_urls(
    xlsx_path: str,
    header_rows: Tuple[int, int] = (3, 5),
    annotate_shaded: bool = True,   # kept for compatibility; we’ll also return separate likert
    n_levels: int = 5,
    theme_override: Dict[int, str] | None = None
) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[2]

    rows = []
    for row in ws.iter_rows(values_only=False):
        row_vals, row_likerts = [], []
        for cell in row:
            val = cell.value
            score = 0.0
            fill = cell.fill
            if fill and fill.patternType:
                try:
                    fgtype = getattr(fill.fgColor, 'type', None)
                    if fgtype == 'rgb' and fill.fgColor.rgb:
                        score = rgb_to_signed_score(fill.fgColor.rgb.upper())
                    elif fgtype == 'theme' and getattr(fill.fgColor, 'theme', None) is not None:
                        rgb = theme_to_rgb(fill.fgColor.theme, theme_override)
                        score = rgb_to_signed_score(rgb)
                    elif fgtype == 'indexed':
                        # could map Excel indexed palette here if needed
                        score = 0.0
                except Exception:
                    pass
            lik = signed_score_to_likert(score, n_levels=n_levels)
            # keep as separate arrays
            row_vals.append(val)
            row_likerts.append(lik)
        rows.append((row_vals, row_likerts))

    # Build DataFrame for values, and another for likert
    df_values = pd.DataFrame([r[0] for r in rows])
    df_likert = pd.DataFrame([r[1] for r in rows])

    # MultiIndex header from df_values (rows are identical positions in both)
    i1, i2 = header_rows[0]-1, header_rows[1]-1
    lvl1, lvl2 = df_values.iloc[i1].tolist(), df_values.iloc[i2].tolist()
    # ffill lvl1
    last = None
    for i, v in enumerate(lvl1):
        if v is None:
            lvl1[i] = last
        else:
            last = v
    tuples = []
    for j, (l1, l2) in enumerate(zip(lvl1, lvl2)):
        L1 = l1 if l1 is not None else f"Unnamed1_{j}"
        L2 = l2 if l2 is not None else f"Unnamed2_{j}"
        tuples.append((str(L1), str(L2)))

    cols = pd.MultiIndex.from_tuples(tuples)
    df_values.columns = cols
    df_likert.columns = pd.MultiIndex.from_tuples([(L1, f"{L2}__likert") for (L1, L2) in tuples])

    # drop header rows
    df_values = df_values.drop(index=range(0, i2+1)).reset_index(drop=True)
    df_likert = df_likert.drop(index=range(0, i2+1)).reset_index(drop=True)

    # Combine value + likert side by side (separate columns, no string concat)
    df = pd.concat([df_values, df_likert], axis=1)

    # --- URL extraction (robust) ---
    # Sort first to keep lexsort invariant
    df = df.sort_index(axis=1)
    lvl2_all = df.columns.get_level_values(1)
    ref_l2 = [c for c in lvl2_all if c is not None and re.search(r'reference\b', str(c), re.I)]
    if ref_l2:
        # pick the first such L2 label
        url_l2 = ref_l2[0]
        # slice all L1 groups having that L2
        ref_sub = df.loc[:, pd.IndexSlice[:, url_l2]].astype(str)
        # merge per row: first non-empty to the right
        ref_ser = (ref_sub.replace({"": pd.NA, "None": pd.NA})
                         .bfill(axis=1)
                         .iloc[:, 0]
                         .fillna(""))
        url_pat = r'(https?://\S+)'
        df[('URLs', 'first_url')] = ref_ser.str.extract(url_pat, expand=False)
        df[('URLs', 'all_urls')]  = ref_ser.apply(lambda s: re.findall(url_pat, s))
        # re-sort after adding new keys
        df = df.sort_index(axis=1)
    else:
        df[('URLs', 'first_url')] = pd.NA
        df[('URLs', 'all_urls')]  = [[] for _ in range(len(df))]
        df = df.sort_index(axis=1)

    return df

def _split_value_likert(s):
    # returns (value_str, likert_int or None)
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return (None, None)
    s = str(s)
    m = re.match(r"^(.*?)(?:\|(-?\d+))?$", s)
    if not m:
        return (s, None)
    val = m.group(1).strip()
    lk  = m.group(2)
    return (val if val != "" else None, (int(lk) if lk is not None else None))

def flatten_df(df):
    # Flatten MultiIndex
    flat_cols = ["|".join([str(c) for c in col if c not in (None, "")])
                 for col in df.columns.to_flat_index()]
    df = df.copy()
    df.columns = flat_cols

    def sget(name):
        return df[name] if name in df.columns else pd.Series([None]*len(df), index=df.index)

    # pick source columns (post header-clean)
    src = {
        "Goal": "Input data|Goal",
        "Target": "Input data|Target",
        "Target description": "Input data|Target description",
        "Dimension of Temporality": "Input data|Dimension of Temporality",
        "Reciprocal Interdependence": "Input data|Reciprocal Interdependence",
        "Justification": "Green Hydrogen Value Chain Justification|COLORED_EMPTY|Justification",
        "Reference": "Green Hydrogen Value Chain Justification|COLORED_EMPTY|Reference",
    }

    out = pd.DataFrame({k: sget(v) for k, v in src.items()})
    # forward fill Goal
    out["Goal"] = out["Goal"].replace(r"^\s*$", None, regex=True).ffill()

    # Split "value|likert" into two columns where applicable
    for k in ["Target", "Target description", "Dimension of Temporality",
              "Reciprocal Interdependence", "Justification", "Reference"]:
        if k in out:
            pairs = out[k].apply(_split_value_likert)
            out[k] = pairs.apply(lambda t: t[0])
            out[f"{k}__likert"] = pairs.apply(lambda t: t[1])

    # Keep rows with ANY useful content
    key_any = ["Target", "Target description", "Justification", "Reference"]
    keep = pd.concat(
        [(out[k].astype(str).str.strip() != "") & out[k].notna() for k in key_any if k in out],
        axis=1
    ).any(axis=1) if any(k in out for k in key_any) else pd.Series(False, index=out.index)
    out = out.loc[keep].copy()

    # URLs from Reference value (not from likert)
    ref_ser = out["Reference"].astype(str) if "Reference" in out else pd.Series([""]*len(out), index=out.index)
    out["extracted_url"] = ref_ser.str.extract(r"(https?://[^\s]+)", expand=False)
    out["all_urls"] = ref_ser.apply(lambda x: re.findall(r"(https?://[^\s]+)", x) if isinstance(x, str) else [])

    return out

def get_pdf_filename_from_url(pdf_url):
    if not pdf_url: return None
    name = os.path.basename(urlparse(pdf_url).path)
    name = unquote(name)
    return name if name.lower().endswith('.pdf') else name + '.pdf'

def download_pdf(pdf_url, save_dir):
    if not pdf_url: return None
    os.makedirs(save_dir, exist_ok=True)
    path = os.path.join(save_dir, get_pdf_filename_from_url(pdf_url))
    try:
        r = requests.get(pdf_url, stream=True)
        if r.status_code==200:
            with open(path,'wb') as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            return path
    except: return None
    return None

# =========================
# Process DataFrame with Elsevier fetching
# =========================
def process_dataframe(df, api_key, save_dir="/content/downloaded_pdfs"):
    df = df.loc[:,~df.columns.duplicated()]
    df['elsevier_abstracts'] = pd.Series([[] for _ in range(len(df))], index=df.index, dtype=object)
    df['other_urls'] = pd.Series([[] for _ in range(len(df))], index=df.index, dtype=object)
    for idx, url_list in df['all_urls'].items():
        if not isinstance(url_list,list) or not url_list: continue
        abstracts, others = [], []
        for url in url_list:
            doi = clean_doi_from_url(url)
            if doi:
                res = fetch_abstract_title_and_pdf(doi, api_key)
                if res:
                    pdf_path = download_pdf(res['pdf_url'], save_dir) if res['pdf_url'] else None
                    abstracts.append({**res,"pdf_local_path":pdf_path})
                else:
                    others.append(url)
            else:
                others.append(url)
        df.at[idx,'elsevier_abstracts']=abstracts
        df.at[idx,'other_urls']=others
    return df

# =========================
# Debug function to inspect DataFrame structure
# =========================
def debug_dataframe_structure(df):
    print("DataFrame shape:", df.shape)
    print("DataFrame columns:")
    for i, col in enumerate(df.columns):
        print(f"  {i}: {col}")
    print("\nFirst few rows:")
    print(df.head())
    return df

# ---- put this near the top of your script (imports) ----
def trigger_download(path: str, *, serve_http: bool = False, http_port: int = 8000) -> None:
    """
    Try to trigger a download depending on the environment.

    - In Google Colab: open the browser 'Save As' dialog.
    - In Jupyter (local): show a clickable link.
    - Otherwise (server / headless):
        * copy to ~/Downloads/<file>
        * print scp / curl / wget commands.
        * optionally start a tiny HTTP server for the directory (serve_http=True).
    """
    path = os.path.abspath(path)
    if not os.path.exists(path):
        print(f"[WARN] File not found: {path}", file=sys.stderr)
        return

    # --- 1) Try Google Colab ---
    try:
        import google.colab.files as gfiles  # type: ignore
        gfiles.download(path)
        print(f"[INFO] Colab download triggered for: {path}")
        return
    except Exception:
        pass

    # --- 2) Try Jupyter / IPython (shows clickable link in notebook) ---
    try:
        from IPython import get_ipython  # type: ignore
        if get_ipython():
            from IPython.display import FileLink, display  # type: ignore
            display(FileLink(path))
            print(f"[INFO] Jupyter link displayed for: {path}")
            return
    except Exception:
        pass

    # --- 3) Streamlit (if you ever run via `streamlit run ...`) ---
    # Note: Rendering a button in a non-streamlit process does nothing, so we only try
    # if a Streamlit runtime is present.
    try:
        import streamlit.runtime  # presence implies running in Streamlit
        import streamlit as st
        with open(path, "rb") as f:
            st.download_button(
                label=f"Download {os.path.basename(path)}",
                data=f,
                file_name=os.path.basename(path),
                mime="text/csv"
            )
        print("[INFO] Streamlit download button rendered.")
        return
    except Exception:
        pass

    # --- 4) Server / headless fallback with home tree access ---
    home = str(Path.home())
    downloads_dir = os.path.join(home, "Downloads")
    os.makedirs(downloads_dir, exist_ok=True)
    out_copy = os.path.join(downloads_dir, os.path.basename(path))

    # Copy if different location
    if os.path.abspath(os.path.dirname(path)) != os.path.abspath(downloads_dir):
        try:
            import shutil
            shutil.copy2(path, out_copy)
            print(f"[INFO] Copied to: {out_copy}")
        except Exception as e:
            print(f"[WARN] Could not copy to ~/Downloads: {e}")

    user = getpass.getuser()
    host = socket.gethostname()
    # Best-effort IP (may be behind NAT; adjust if needed)
    try:
        ip = socket.gethostbyname(host)
    except Exception:
        ip = "SERVER_IP"

    print("\n[HOW TO DOWNLOAD THIS FILE]")
    print("1) From your local machine (in a terminal), using scp:")
    print(f"   scp {user}@{host}:{path} ./")
    print("\n2) Or with scp (copy from ~/Downloads):")
    print(f"   scp {user}@{host}:{out_copy} ./")
    print("\n3) Or with curl/wget (if reachable):")
    print(f"   curl -O http://{ip}:{http_port}/{os.path.basename(path)}")
    print(f"   wget http://{ip}:{http_port}/{os.path.basename(path)}")

    if serve_http:
        # Serve the directory containing the file (blocks this process)
        from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
        serve_dir = os.path.dirname(path)
        os.chdir(serve_dir)
        print(f"\n[INFO] Serving {serve_dir} at http://0.0.0.0:{http_port}/")
        print("      Press Ctrl+C to stop.")
        try:
            ThreadingHTTPServer(("0.0.0.0", http_port), SimpleHTTPRequestHandler).serve_forever()
        except KeyboardInterrupt:
            print("\n[INFO] HTTP server stopped.")


# =========================
# Main CLI
# =========================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", help="Input Excel .xlsx file (OpenXML)")
    parser.add_argument("output_dir", help="Output directory for the CSV")
    parser.add_argument("--api_key", help="Elsevier API key", required=True)
    parser.add_argument("--debug", action="store_true", help="Enable debug output")
    args = parser.parse_args()

    try:
        # 1) Extract + URL columns from the specified sheet
        df = extract_third_sheet_multiheader_and_urls(args.xlsx_path)
        if args.debug:
            print("=== Debug: After extract_sheet_and_urls ===")
            debug_dataframe_structure(df)

        # 2) Flatten to single-level columns and pick relevant fields
        flat_df = flatten_df(df)
        if args.debug:
            print("=== Debug: After flatten_df ===")
            debug_dataframe_structure(flat_df)

        # 3) Enrich with Elsevier metadata + optional PDF download
        flat_df_processed = process_dataframe(flat_df, args.api_key)
        if args.debug:
            print("=== Debug: After process_dataframe ===")
            print("Rows:", len(flat_df_processed))
            if "all_urls" in flat_df_processed.columns:
                print("Rows with any URL:",
                      sum(bool(x) for x in flat_df_processed["all_urls"]))

        # 4) Save exactly once (final artifact), then trigger download
        out_path = Path(args.output_dir) / "flattened_with_elsevier.csv"
        out_path.parent.mkdir(parents=True, exist_ok=True)
        flat_df_processed.to_csv(out_path, index=False)
        print(f"Saved: {out_path}")

        # 5) Download / fallback copy & instructions
        trigger_download(str(out_path), serve_http=False)

    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        if args.debug:
            import traceback
            traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
"""
Created on Fri Sep 19 17:58:25 2025

@author: niran
"""