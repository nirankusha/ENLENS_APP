# -*- coding: utf-8 -*-
#!/usr/bin/env python3
import argparse
import os
import re
import requests
import pandas as pd
import colorsys
from openpyxl import load_workbook
from urllib.parse import urlparse, unquote
import ast

# =========================
# Color / ordinal helpers
# =========================
def rgb_to_signed_score(rgb_hex):
    if rgb_hex is None:
        return 0
    if len(rgb_hex) == 8:
        rgb_hex = rgb_hex[2:]
    r = int(rgb_hex[0:2], 16) / 255
    g = int(rgb_hex[2:4], 16) / 255
    b = int(rgb_hex[4:6], 16) / 255
    h, s, v = colorsys.rgb_to_hsv(r, g, b)
    hue_deg = h * 360
    intensity = s * v
    if 90 <= hue_deg <= 150:    # green
        return intensity
    elif hue_deg <= 15 or hue_deg >= 345:  # red
        return -intensity
    return 0

def theme_to_rgb(theme_index):
    theme_map = {9: 'FF00FF00', 10: 'FFFF0000', 0: None}
    return theme_map.get(theme_index, None)

def signed_score_to_likert(score, n_levels=5):
    if score == 0:
        return 0
    abs_score = min(abs(score), 1.0)
    level = max(1, round(abs_score * n_levels))
    return level if score > 0 else -level

# =========================
# Excel extraction
# =========================
def extract_sheet_and_urls(xlsx_path, header_rows=(3,5), annotate_shaded=True, n_levels=5):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.worksheets[2]
    all_rows = []

    for row in ws.iter_rows(values_only=False):
        row_out = []
        for cell in row:
            val = cell.value
            score = 0
            fill = cell.fill

            if fill and fill.patternType:
                t = getattr(fill.fgColor, 'type', None)
                if t == 'rgb' and fill.fgColor.rgb:
                    score = rgb_to_signed_score(fill.fgColor.rgb.upper())
                elif t == 'theme' and fill.fgColor.theme is not None:
                    rgb = theme_to_rgb(fill.fgColor.theme)
                    score = rgb_to_signed_score(rgb)

            likert = signed_score_to_likert(score, n_levels)
            cell_out = f"{val}|{likert}" if (val is not None and annotate_shaded) else val or likert
            row_out.append(cell_out)
        all_rows.append(row_out)

    df = pd.DataFrame(all_rows)

    # MultiIndex header
    idx1, idx2 = header_rows[0]-1, header_rows[1]-1
    lvl1, lvl2 = df.iloc[idx1].tolist(), df.iloc[idx2].tolist()
    filled_lvl1 = []
    last = None
    for v in lvl1:
        filled_lvl1.append(v if v is not None else last)
        if v is not None: last = v
    tuples = [(l1 if l1 is not None else f"U1_{i}", l2 if l2 is not None else f"U2_{i}") for i,(l1,l2) in enumerate(zip(filled_lvl1,lvl2))]
    df.columns = pd.MultiIndex.from_tuples(tuples)

    # Drop header rows
    df = df.drop(index=range(0, idx2+1)).reset_index(drop=True)

    # Extract URLs from "Reference" column
    url_col = [c for c in df.columns.get_level_values(1) if "Reference" in str(c)][0]
    ref_ser = df.xs(url_col, axis=1, level=1).squeeze()
    df[('URLs', 'first_url')] = ref_ser.str.extract(r'(https?://[^\s]+)', expand=False)
    df[('URLs', 'all_urls')] = ref_ser.apply(lambda x: re.findall(r'(https?://[^\s]+)', str(x)))

    return df

def flatten_df(df):
    df.columns = ["|".join([str(c) for c in col if c not in (None,"")]) for col in df.columns.to_flat_index()]
    flat_df = pd.DataFrame({
        "Goal": df.get("Input data|Goal"),
        "Target": df.get("Input data|Target"),
        "Target description": df.get("Input data|Target description"),
        "Dimension of Temporality": df.get("Input data|Dimension of Temporality"),
        "Reciprocal Interdependence": df.get("Input data|Reciprocal Interdependence"),
        "Justification": df.get("Green Hydrogen Value Chain Justification|COLORED_EMPTY|Justification"),
        "Reference": df.get("Green Hydrogen Value Chain Justification|COLORED_EMPTY|Reference"),
    })
    flat_df["Goal"] = flat_df["Goal"].ffill()
    flat_df = flat_df.replace(r'^\s*$', None, regex=True)
    flat_df = flat_df.dropna(subset=["Target"], how="all")
    flat_df['extracted_url'] = flat_df['Reference'].str.extract(r'(https?://[^\s]+)', expand=False)
    flat_df['all_urls'] = flat_df['Reference'].apply(lambda x: re.findall(r'(https?://[^\s]+)', str(x)))
    return flat_df

# =========================
# Elsevier API helpers
# =========================
def clean_doi_from_url(url): 
    m = re.search(r'10\.\d{4,9}/[-._;()/:A-Z0-9]+', url, re.I)
    return m.group(0).rstrip('.') if m else None

def fetch_abstract_title_and_pdf(doi, api_key):
    if not api_key:
        return None
    if not doi.lower().startswith("10.1016/"):
        return None
    headers = {"X-ELS-APIKey": api_key, "Accept":"application/json"}
    url = f"https://api.elsevier.com/content/article/doi/{doi}?view=META_ABS"
    try:
        r = requests.get(url, headers=headers)
        if r.status_code != 200: return None
        data = r.json()
        article = data.get('full-text-retrieval-response',{})
        core = article.get('coredata',{})
        links = article.get('link',[])
        if isinstance(links, dict): links=[links]
        pdf_url = next((l.get('@href') for l in links if l.get('@ref') in ('pdf','full-text')), None)
        return {"title":core.get('dc:title','No title'),
                "abstract":core.get('dc:description','No abstract'),
                "pdf_url": pdf_url}
    except: return None

def get_pdf_filename_from_url(pdf_url):
    if not pdf_url: return None
    name = os.path.basename(urlparse(pdf_url).path)
    name = unquote(name)
    return name if name.lower().endswith('.pdf') else name + '.pdf'

def download_pdf(pdf_url, save_dir):
    if not pdf_url: return None
    os.makedirs(save_dir, exist_ok=True)
    path = os.path.join(save_dir, get_pdf_filename_from_url(pdf_url))
    try:
        r = requests.get(pdf_url, stream=True)
        if r.status_code==200:
            with open(path,'wb') as f:
                for chunk in r.iter_content(8192):
                    f.write(chunk)
            return path
    except: return None
    return None

# =========================
# Process DataFrame with Elsevier fetching
# =========================
def process_dataframe(df, api_key, save_dir="/content/downloaded_pdfs"):
    df = df.loc[:,~df.columns.duplicated()]
    df['elsevier_abstracts'] = pd.Series([[] for _ in range(len(df))], index=df.index, dtype=object)
    df['other_urls'] = pd.Series([[] for _ in range(len(df))], index=df.index, dtype=object)
    for idx, url_list in df['all_urls'].items():
        if not isinstance(url_list,list) or not url_list: continue
        abstracts, others = [], []
        for url in url_list:
            doi = clean_doi_from_url(url)
            if doi:
                res = fetch_abstract_title_and_pdf(doi, api_key)
                if res:
                    pdf_path = download_pdf(res['pdf_url'], save_dir) if res['pdf_url'] else None
                    abstracts.append({**res,"pdf_local_path":pdf_path})
                else:
                    others.append(url)
            else:
                others.append(url)
        df.at[idx,'elsevier_abstracts']=abstracts
        df.at[idx,'other_urls']=others
    return df

# =========================
# Main CLI
# =========================
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", help="Input Excel file")
    parser.add_argument("output_dir", help="Output CSV directory")
    parser.add_argument("--api_key", help="Elsevier API key", required=True)
    args = parser.parse_args()

    df = extract_sheet_and_urls(args.xlsx_path)
    flat_df = flatten_df(df)
    flat_df_processed = process_dataframe(flat_df, args.api_key)

    os.makedirs(args.output_dir, exist_ok=True)
    out_path = os.path.join(args.output_dir, "flattened_with_elsevier.csv")
    flat_df_processed.to_csv(out_path, index=False)
    print(f"Saved: {out_path}")
    print(flat_df_processed.head())

if __name__ == "__main__":
    main()

"""
Created on Fri Sep 19 17:58:25 2025

@author: niran
"""

